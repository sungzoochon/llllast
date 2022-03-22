from datetime import datetime
from openpyxl import load_workbook
import datetime
import pyupbit
access = "dNlYFPozFWnpa6gtT7OuOjLFbx7iD65qCkWVxxTg"
secret = "nzhg0ah7Lh3XX2nKlUxBY904RBoSbpBHMrf8pGrZ"
upbit = pyupbit.Upbit(access, secret)
wb = load_workbook('수익률 계산.xlsx')
ws = wb.active
def get_balance(ticker):
    """잔고 조회"""
    balances = upbit.get_balances()
    for b in balances:
        if b['currency'] == ticker:
            if b['balance'] is not None:
                return float(b['balance'])
            else:
                return 0
    return 0
def KRW_record(KRW,date,original_amount,a):
 ws.cell(1,a + 1,value=date)
 ws.cell(2,a + 1,value=KRW/original_amount)
 ws.cell(3,1,value = a + 1)
 wb.save("수익률 계산.xlsx")
original_amount = 123000
a = int(ws['A3'].value) 
btc = get_balance("KRW")
date = datetime.datetime.now().date()
KRW_record(btc,date,original_amount,a)


